from openpyxl import load_workbook
import numpy as np
import re
import jieba
import os
from gensim.models import KeyedVectors
import tensorflow as tf
from tensorflow.python.keras.models import Sequential
from tensorflow.python.keras.layers import Dense, Embedding, LSTM, Bidirectional
from tensorflow.python.keras.preprocessing.sequence import pad_sequences
from tensorflow.python.keras.callbacks import EarlyStopping, ModelCheckpoint, ReduceLROnPlateau
from sklearn.model_selection import train_test_split
import warnings
warnings.filterwarnings("ignore")

# 加载word2vec预训练模型
cn_model = KeyedVectors.load_word2vec_format(
    './pretrain_models/sgns.zhihu.bigram', binary=False, unicode_errors="ignore"
)


def LoadData(path):
    print('开始加载数据...')
    train_texts_orig = []  # 评价
    train_target = []  # 标签(正负面)
    wb = load_workbook(path)
    ws = wb['宝马']
    for i in range(2, ws.max_row + 1):  # cell的row,column从1开始
        train_texts_orig.append(ws.cell(row=i, column=4).value)
        train_target.append(ws.cell(row=i, column=3).value)

    print('正面数据有 {} 个.'.format(list(train_target).count(1)))
    print('负面数据有 {} 个.'.format(list(train_target).count(0)))

    return train_texts_orig, train_target


def TrainModel(train_texts_orig, train_target):
    print('开始分词等预处理...')
    embedding_dim = 300
    train_tokens = []
    for text in train_texts_orig:
        text = re.sub("[\\s+\\.\\!\\/_,$%^*(+\"\']+|[+——！，。？、~@#￥%……&*（）]+", "", text)  # 去掉标点
        cut = jieba.cut(text)  # 结巴分词
        cut_list = [i for i in cut]  # 分词结果转换为list
        for i, word in enumerate(cut_list):
            try:
                cut_list[i] = cn_model.key_to_index[word]  # 将词转换为词向量的索引
            except KeyError:
                cut_list[i] = 0  # 如果词不在字典中，则输出0
        train_tokens.append(cut_list)

    num_tokens = [len(token) for token in train_tokens]  # 每条评论分词后 词的个数
    num_tokens = np.array(num_tokens)
    max_tokens = np.mean(num_tokens) + 2 * np.std(num_tokens)
    max_tokens = int(max_tokens)
    print('max_tokens = {}'.format(max_tokens))

    num_words = 60000  # 最多60000个词(向量)
    embedding_matrix = np.zeros((num_words, embedding_dim))  # 初始化embedding_matrix
    for i in range(num_words):
        embedding_matrix[i, :] = cn_model[cn_model.index_to_key[i]]
    embedding_matrix = embedding_matrix.astype('float32')

    # padding
    train_pad = pad_sequences(train_tokens, maxlen=max_tokens, padding='pre', truncating='pre')
    train_pad[train_pad >= num_words] = 0  # 超过的词用0代替

    train_target = np.array(train_target)
    train_target = train_target.astype('int')

    print('搭建模型...')
    # 90%的样本用来训练，剩余10%用来测试
    X_train, X_test, y_train, y_test = train_test_split(train_pad,
                                                        train_target,
                                                        test_size=0.1,
                                                        random_state=12)

    model = Sequential()
    model.call = tf.function(model.call)
    model.add(Embedding(num_words,
                        embedding_dim,
                        weights=[embedding_matrix],
                        input_length=max_tokens,
                        trainable=False))
    model.add(Bidirectional(LSTM(units=64, return_sequences=True)))
    model.add(LSTM(units=16, return_sequences=False))
    model.add(Dense(1, activation='sigmoid'))

    print(model.summary())  # 输出模型信息

    model.compile(loss='binary_crossentropy',
                  optimizer=tf.keras.optimizers.SGD(lr=1e-3),
                  metrics=['accuracy'])

    # 建立权重的存储点
    path_checkpoint = './pretrain_models/sentiment_checkpoint.keras'
    checkpoint = ModelCheckpoint(filepath=path_checkpoint, monitor='val_loss',
                                 verbose=1, save_weights_only=True,
                                 save_best_only=True)

    # 尝试加载已训练模型
    try:
        model.load_weights(path_checkpoint)
    except Exception as e:
        print(e)

    # 如果3个epoch内validation loss没有改善则停止训练
    earlystopping = EarlyStopping(monitor='val_loss', patience=5, verbose=1)

    # 训练时自动调整learning rate
    lr_reduction = ReduceLROnPlateau(monitor='val_loss', factor=0.1, min_lr=1e-8, patience=0, verbose=1)

    # 定义callback函数
    callbacks = [earlystopping, checkpoint, lr_reduction]

    # 开始训练
    print('开始训练模型...')
    model.fit(X_train, y_train, validation_split=0.1, epochs=40, batch_size=256, callbacks=callbacks)

    # 模型保存
    model.save('./pretrain_models/sentiment_model.h5')

    return model, max_tokens


def PredictSentiment(text):
    print('当前测试的文本是: {}'.format(text))
    max_tokens = 49
    # 去标点
    text = re.sub("[\\s+\\.\\!\\/_,$%^*(+\"\']+|[+——！，。？、~@#￥%……&*（）]+", "", text)
    # 分词
    cut = jieba.cut(text)
    cut_list = [i for i in cut]
    # tokenize
    for i, word in enumerate(cut_list):
        try:
            cut_list[i] = cn_model.key_to_index[word]
            if cut_list[i] >= 30000:
                cut_list[i] = 0
        except KeyError:
            cut_list[i] = 0
    # padding
    tokens_pad = pad_sequences([cut_list], maxlen=max_tokens, padding='pre', truncating='pre')
    # 预测
    model = tf.keras.models.load_model('./pretrain_models/sentiment_model.h5')  # 加载模型
    result = model.predict(x=tokens_pad)
    coef = result[0][0]
    if coef >= 0.5:
        print('是一例正面评价', 'output=%.2f' % coef)
        return '正面'
    else:
        print('是一例负面评价', 'output=%.2f' % coef)
        return '负面'


if __name__ == '__main__':
    model_dir = './pretrain_models/'
    if 'sentiment_model.h5' not in os.listdir(model_dir):
        train_texts_orig, train_target = LoadData('./data/autohome_sentiment.xlsx')
        model, max_tokens = TrainModel(train_texts_orig, train_target)

    text = '品控不好，还没到一个月就坏了'
    PredictSentiment(text)
